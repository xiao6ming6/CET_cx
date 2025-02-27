#!/usr/bin/python3
# -*- coding:utf-8 -*-
# @author  : xiumi
# @time    : 2024/12/13 16:12
# @function: the script is used to do ...
# @version : V1
import os
import time
import pandas as pd
from openpyxl.reader.excel import load_workbook
from selenium.common import TimeoutException
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook

def get_info(user_name, id_code, bmkm):



    # 设置 Chrome 浏览器选项
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")  # 禁用 GPU 加速
    chrome_options.add_argument("--no-sandbox")  # 避免一些浏览器启动错误

    # 启动浏览器
    driver = uc.Chrome(options=chrome_options, driver_executable_path=ChromeDriverManager().install())

    # 访问页面
    driver.get("https://cjcx.neea.edu.cn/html1/folder/21083/9970-1.htm")

    # 显式等待，直到 id 为 selProvince 的元素加载完成
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "km")))

    # 选择 selProvince 下拉框中的 value=41
    province_select = Select(driver.find_element(By.ID, "km"))
    province_select.select_by_value(bmkm)

    # 在 txtIDNumber 输入框中输入 idCode
    driver.find_element(By.ID, "no").send_keys(id_code)

    # 在 txtName 输入框中输入 user_name
    driver.find_element(By.ID, "xm").send_keys(user_name)

    # 点击查询
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "submitButton")))
    driver.find_element(By.ID, "submitButton").click()
    # # 提示用户手动输入验证码，等待验证码输入框
    # print("请输入验证码，点击 '登录' 按钮后，继续执行程序")
    # WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "txtVerificationCode")))
    #
    # # 用户手动输入验证码，暂停执行直到输入完成
    # while not driver.find_element(By.ID, "txtVerificationCode").get_attribute("value"):
    #     time.sleep(1)  # 每秒检查一次是否有输入
    #
    # # 获取用户输入的验证码
    # verification_code = driver.find_element(By.ID, "txtVerificationCode").get_attribute("value")
    #
    # # 在 txtVerificationCode 输入框中输入用户输入的验证码
    # driver.find_element(By.ID, "txtVerificationCode").send_keys(verification_code)



    # 检测直到登录按钮消失或不可点击
    try:
        WebDriverWait(driver, 3).until(EC.invisibility_of_element_located((By.ID, "submitButton")))
    except TimeoutException:
        # print('wait_timeout')
        try:
            # driver.find_element(By.ID, "submitButton").click()
            # WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "layui-layer-content")))
            # # 如果找到了 'messager-error' 元素，表示不对
            # massage = driver.find_element(By.ID, "layui-layer-content").text
            # print(massage)  # 登录按钮不可用或已消失，继续执行后续操作
            if bmkm == 2:
                return get_info(user_name, id_code, 1)
            else:
                # with open("data/cj_cet_data.txt", "a", encoding="utf-8") as file:
                #     file.write(f"{user_name}\t未参加\n")
                wb = load_workbook('data/cj_cet_data.xlsx')
                ws = wb.active
                t_score = [user_name, '未参加']
                ws.append(t_score)
                wb.save('data/cj_cet_data.xlsx')
                driver.quit()  # 关闭浏览器
                return  # 退出函数，不继续执行后续操作

        except TimeoutException:
            print("未找到 'messager-error' 元素，继续执行后续操作")
            return

    # 等待页面加载
    WebDriverWait(driver, 6).until(EC.invisibility_of_element_located((By.ID, "score")))

    # 分数的内容
    score = driver.find_element(By.CSS_SELECTOR, ".cjd-bot-txt strong[code='score']").text # 总分
    sco_lc = driver.find_element(By.CSS_SELECTOR, ".cjd-bot-txt i[code='sco_lc']").text # 听力
    sco_rd = driver.find_element(By.CSS_SELECTOR, ".cjd-bot-txt i[code='sco_rd']").text # 阅读
    sco_wt = driver.find_element(By.CSS_SELECTOR, ".cjd-bot-txt i[code='sco_wt']").text # 写作

    # if bmkm == 1:
    #     level = '四级'
    # elif bmkm == 2:
    #     level = '六级'
    # else:
    #     print(f'{user_name}:error')
    level = '四级'
    if bmkm == 2 or bmkm == '2': level = '六级'

    # t_score = f'{level}\t{score}\t{sco_lc}\t{sco_rd}\t{sco_wt}'
    #
    #
    # # 将数据写入 cet_data.txt 文件
    # with open("data/cj_cet_data.txt", "a", encoding="utf-8") as file:
    #     file.write(f"{user_name}\t{t_score}\n")
    print(f'{level}\t{score}\t{sco_lc}\t{sco_rd}\t{sco_wt}')

    # 如果文件不存在，则创建一个新的 Workbook
    cj_file_path = 'data/cj_cet_data.xlsx'
    if not os.path.exists(cj_file_path):
        wb = Workbook()
        ws = wb.active
        # 给表格添加表头
        ws.append(["姓名", "科目", "总分", "听力", "阅读", "写作"])
    else:# 如果文件存在，则打开现有文件
        wb = load_workbook(cj_file_path)
        ws = wb.active

    t_score = [user_name, level, score, sco_lc, sco_rd, sco_wt]
    ws.append(t_score)
    wb.save(cj_file_path)

    # 关闭浏览器
    driver.quit()



def get_km():
    default_subjects = ["英语四级笔试", "英语六级笔试"]
    # 存储所有人的报名科目
    user_subjects = {}

    # 读取 cet_data.txt 文件
    with open('data/cet_data.txt', 'r', encoding='utf-8') as file:
        for line in file:
            # 去除行尾的换行符并分割为名字和科目
            line = line.strip()
            name, subject = line.split("\t")
            name = name.strip()
            subject = subject.strip()

            # 如果科目是“未报名或不可查”，则插入两条默认数据
            if subject == "未报名或不可查":
                user_subjects[name] = '2'
            else:
                # 否则直接保存科目信息
                if subject == default_subjects[0]:
                    user_subjects[name] = '1'
                elif subject in default_subjects[1]:
                    user_subjects[name] = '2'
    return user_subjects

# name = '李明洋'
# idCard = '411525200309016616'
# get_info(name, idCard)


name_dict = {}
# 读取 Excel 文件
file_path = 'data/data.xlsx'
try:
    df = pd.read_excel(file_path)
    index = 0
    # 遍历每一行数据
    for _, row in df.iterrows():
        user_name = row['姓名']
        id_card = row['身份证号']
        index +=1

        name_dict[user_name] = id_card
    print(f'总数：{index}')
except Exception as e:
    print(f"Error reading the Excel file: {e}")

km = get_km()

for key,value in name_dict.items():
    bmkm = km[key]
    print(f"{key}\t{value}\t{bmkm}")
    get_info(key, value, bmkm)